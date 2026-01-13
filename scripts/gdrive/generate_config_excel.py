#!/usr/bin/env python3
"""
Genera file Excel con tutte le variabili di configurazione MeepleAI.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path

# Colori
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
CATEGORY_FILL = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
CATEGORY_FONT = Font(bold=True, size=11)
REQUIRED_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
OPTIONAL_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
HIGH_SECURITY_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Dati delle variabili
CONFIG_DATA = [
    # Categoria, Variabile, File Sorgente, Posizione, Esempio, Obbligatorio, Sicurezza, Note

    # === DATABASE POSTGRESQL ===
    ("Database PostgreSQL", "POSTGRES_USER", ".env.local", "~10", "postgres", "✅ Sì", "🟢 Basso", "Username database"),
    ("Database PostgreSQL", "POSTGRES_PASSWORD", ".env.local", "~11", "meeplepass", "✅ Sì", "🔴 Alto", "⚠️ Cambiare in produzione!"),
    ("Database PostgreSQL", "POSTGRES_DB", ".env.local", "~12", "meepleai", "✅ Sì", "🟢 Basso", "Nome database"),
    ("Database PostgreSQL", "POSTGRES_HOST", ".env.local", "~13", "postgres (Docker) / localhost", "✅ Sì", "🟢 Basso", "Hostname database"),
    ("Database PostgreSQL", "POSTGRES_PORT", ".env.local", "~14", "5432", "✅ Sì", "🟢 Basso", "Porta standard PostgreSQL"),
    ("Database PostgreSQL", "ConnectionStrings__Postgres", "appsettings.json", "ConnectionStrings", "Host=postgres;Database=meepleai;Username=postgres;Password=meeplepass;Pooling=true;Minimum Pool Size=2;Maximum Pool Size=20", "✅ Sì", "🔴 Alto", "Connection string completa"),

    # === REDIS CACHE ===
    ("Redis Cache", "REDIS_URL", ".env.local", "~20", "redis:6379", "✅ Sì", "🟢 Basso", "URL Redis"),
    ("Redis Cache", "REDIS_PASSWORD", ".env.local", "~21", "(vuoto in dev)", "❌ Dev / ✅ Prod", "🔴 Alto", "Password Redis (vuota in dev)"),
    ("Redis Cache", "HYBRIDCACHE_ENABLE_L2", ".env.local", "~22", "false (dev) / true (prod)", "❌ No", "🟢 Basso", "Abilita cache L2 distribuita"),

    # === QDRANT VECTOR DB ===
    ("Qdrant Vector DB", "QDRANT_URL", ".env.local", "~25", "http://qdrant:6333", "✅ Sì", "🟢 Basso", "URL API Qdrant"),
    ("Qdrant Vector DB", "QDRANT_API_KEY", ".env.local", "~26", "(vuoto in dev)", "❌ Dev / ✅ Prod", "🟡 Medio", "API Key Qdrant"),

    # === ASP.NET CORE API ===
    ("ASP.NET Core API", "ASPNETCORE_ENVIRONMENT", ".env.local", "~30", "Development / Production", "✅ Sì", "🟢 Basso", "Ambiente runtime"),
    ("ASP.NET Core API", "ASPNETCORE_URLS", ".env.local", "~31", "http://+:8080", "✅ Sì", "🟢 Basso", "URL binding API"),
    ("ASP.NET Core API", "ASPNETCORE_DETAILEDERRORS", ".env.local", "~32", "true (dev) / false (prod)", "❌ No", "🟢 Basso", "Mostra errori dettagliati"),
    ("ASP.NET Core API", "JWT_ISSUER", ".env.local", "~35", "http://localhost:8080", "✅ Sì", "🟢 Basso", "Issuer JWT token"),
    ("ASP.NET Core API", "JWT_AUDIENCE", ".env.local", "~36", "http://localhost:3000", "✅ Sì", "🟢 Basso", "Audience JWT token"),
    ("ASP.NET Core API", "ALLOW_ORIGIN", ".env.local", "~37", "http://localhost:3000", "✅ Sì", "🟢 Basso", "CORS origin consentito"),

    # === NEXT.JS FRONTEND ===
    ("Next.js Frontend", "NODE_ENV", ".env.local", "~45", "development / production", "✅ Sì", "🟢 Basso", "Ambiente Node.js"),
    ("Next.js Frontend", "NEXT_PUBLIC_API_BASE", ".env.local", "~46", "http://localhost:8080", "✅ Sì", "🟢 Basso", "URL base API backend"),
    ("Next.js Frontend", "NEXT_PUBLIC_TENANT_ID", ".env.local", "~47", "dev", "✅ Sì", "🟢 Basso", "ID tenant multi-tenant"),
    ("Next.js Frontend", "NEXT_PUBLIC_SITE_URL", ".env.local", "~48", "http://localhost:3000", "✅ Sì", "🟢 Basso", "URL pubblico sito"),
    ("Next.js Frontend", "NEXT_TELEMETRY_DISABLED", ".env.local", "~50", "1", "❌ No", "🟢 Basso", "Disabilita telemetria Next.js"),
    ("Next.js Frontend", "NEXT_PUBLIC_LOG_ENDPOINT", ".env.local", "~51", "http://localhost:8080/api/v1/logs", "❌ No", "🟢 Basso", "Endpoint logging client-side"),
    ("Next.js Frontend", "NEXT_PUBLIC_SENTRY_DSN", ".env.local", "~52", "(vuoto in dev)", "❌ No", "🟡 Medio", "DSN Sentry per error tracking"),
    ("Next.js Frontend", "NEXT_PUBLIC_RETRY_ENABLED", ".env.local", "~55", "true", "❌ No", "🟢 Basso", "Abilita retry automatico"),
    ("Next.js Frontend", "NEXT_PUBLIC_RETRY_MAX_ATTEMPTS", ".env.local", "~56", "2", "❌ No", "🟢 Basso", "Max tentativi retry"),

    # === AI/ML SERVICES ===
    ("AI/ML Services", "EMBEDDING_PROVIDER", ".env.local", "~60", "ollama", "✅ Sì", "🟢 Basso", "Provider embeddings (ollama/openai)"),
    ("AI/ML Services", "OLLAMA_URL", ".env.local", "~61", "http://ollama:11434", "✅ Sì", "🟢 Basso", "URL servizio Ollama"),
    ("AI/ML Services", "EMBEDDING_MODEL", ".env.local", "~62", "nomic-embed-text", "✅ Sì", "🟢 Basso", "Modello embedding"),
    ("AI/ML Services", "LOCAL_EMBEDDING_URL", ".env.local", "~63", "http://embedding-service:8000", "✅ Sì", "🟢 Basso", "URL embedding service locale"),
    ("AI/ML Services", "EMBEDDING_FALLBACK_ENABLED", ".env.local", "~64", "true", "❌ No", "🟢 Basso", "Abilita fallback embeddings"),
    ("AI/ML Services", "OPENROUTER_API_KEY", ".env.local", "~66", "sk-or-v1-xxxxx", "✅ Sì", "🔴 Alto", "⚠️ ESSENZIALE per AI chat! Ottieni da openrouter.ai"),
    ("AI/ML Services", "OPENAI_API_KEY", ".env.local", "~67", "sk-xxxxx", "❌ No", "🔴 Alto", "Opzionale: API key OpenAI diretta"),

    # === PDF PROCESSING ===
    ("PDF Processing", "UNSTRUCTURED_STRATEGY", ".env.local", "~75", "fast / hi_res", "✅ Sì", "🟢 Basso", "Strategia estrazione (fast=veloce, hi_res=qualità)"),
    ("PDF Processing", "LANGUAGE", ".env.local", "~76", "ita / eng", "✅ Sì", "🟢 Basso", "Lingua documento"),
    ("PDF Processing", "MAX_FILE_SIZE", ".env.local", "~77", "52428800", "❌ No", "🟢 Basso", "Max dimensione file (50MB)"),
    ("PDF Processing", "TIMEOUT", ".env.local", "~78", "30", "❌ No", "🟢 Basso", "Timeout estrazione (secondi)"),
    ("PDF Processing", "QUALITY_THRESHOLD", ".env.local", "~79", "0.75", "❌ No", "🟢 Basso", "Soglia qualità estrazione"),
    ("PDF Processing", "MIN_CHARS_PER_PAGE", ".env.local", "~80", "400", "❌ No", "🟢 Basso", "Min caratteri per pagina valida"),
    ("PDF Processing", "DEVICE", ".env.local", "~82", "cpu / cuda", "✅ Sì", "🟢 Basso", "Device per ML (cpu/cuda)"),
    ("PDF Processing", "MODEL_NAME", ".env.local", "~83", "docling-project/SmolDocling-256M-preview", "✅ Sì", "🟢 Basso", "Modello VLM per estrazione"),
    ("PDF Processing", "TORCH_DTYPE", ".env.local", "~84", "float32 / bfloat16", "❌ No", "🟢 Basso", "Tipo dati PyTorch"),
    ("PDF Processing", "IMAGE_DPI", ".env.local", "~85", "150", "❌ No", "🟢 Basso", "DPI rendering immagini"),
    ("PDF Processing", "MAX_PAGES_PER_REQUEST", ".env.local", "~86", "10", "❌ No", "🟢 Basso", "Max pagine per richiesta"),

    # === OAUTH AUTHENTICATION ===
    ("OAuth Authentication", "GOOGLE_OAUTH_CLIENT_ID", ".env.local", "~95", "xxxxx.apps.googleusercontent.com", "❌ Dev / ✅ Prod", "🟡 Medio", "Client ID Google OAuth"),
    ("OAuth Authentication", "GOOGLE_OAUTH_CLIENT_SECRET", ".env.local", "~96", "GOCSPX-xxxxx", "❌ Dev / ✅ Prod", "🔴 Alto", "Client Secret Google OAuth"),
    ("OAuth Authentication", "DISCORD_OAUTH_CLIENT_ID", ".env.local", "~98", "123456789012345678", "❌ No", "🟡 Medio", "Client ID Discord OAuth"),
    ("OAuth Authentication", "DISCORD_OAUTH_CLIENT_SECRET", ".env.local", "~99", "xxxxx", "❌ No", "🔴 Alto", "Client Secret Discord OAuth"),
    ("OAuth Authentication", "GITHUB_OAUTH_CLIENT_ID", ".env.local", "~101", "Iv1.xxxxx", "❌ No", "🟡 Medio", "Client ID GitHub OAuth"),
    ("OAuth Authentication", "GITHUB_OAUTH_CLIENT_SECRET", ".env.local", "~102", "xxxxx", "❌ No", "🔴 Alto", "Client Secret GitHub OAuth"),

    # === ADMIN & SECURITY ===
    ("Admin & Security", "INITIAL_ADMIN_EMAIL", ".env.local", "~110", "admin@meepleai.dev", "✅ Sì", "🟢 Basso", "Email admin iniziale"),
    ("Admin & Security", "INITIAL_ADMIN_PASSWORD", ".env.local", "~111", "Demo123!", "✅ Sì", "🔴 Alto", "⚠️ Cambiare in produzione!"),
    ("Admin & Security", "INITIAL_ADMIN_DISPLAY_NAME", ".env.local", "~112", "Local Admin", "❌ No", "🟢 Basso", "Nome visualizzato admin"),
    ("Admin & Security", "SESSION_EXPIRATION_DAYS", ".env.local", "~115", "30", "❌ No", "🟢 Basso", "Durata sessione (giorni)"),
    ("Admin & Security", "INACTIVITY_TIMEOUT_DAYS", ".env.local", "~116", "7", "❌ No", "🟢 Basso", "Timeout inattività (giorni)"),

    # === N8N WORKFLOW ===
    ("n8n Workflow", "N8N_HOST", ".env.local", "~125", "localhost", "✅ Sì", "🟢 Basso", "Host n8n"),
    ("n8n Workflow", "N8N_PORT", ".env.local", "~126", "5678", "✅ Sì", "🟢 Basso", "Porta n8n"),
    ("n8n Workflow", "N8N_PROTOCOL", ".env.local", "~127", "http / https", "✅ Sì", "🟢 Basso", "Protocollo n8n"),
    ("n8n Workflow", "WEBHOOK_URL", ".env.local", "~128", "http://localhost:5678/", "✅ Sì", "🟢 Basso", "URL webhook n8n"),
    ("n8n Workflow", "N8N_BASIC_AUTH_ACTIVE", ".env.local", "~130", "true", "✅ Sì", "🟢 Basso", "Abilita auth base n8n"),
    ("n8n Workflow", "N8N_BASIC_AUTH_USER", ".env.local", "~131", "admin", "✅ Sì", "🟡 Medio", "Username auth n8n"),
    ("n8n Workflow", "N8N_BASIC_AUTH_PASSWORD", ".env.local", "~132", "admin", "✅ Sì", "🔴 Alto", "⚠️ Cambiare in produzione!"),
    ("n8n Workflow", "N8N_ENCRYPTION_KEY", ".env.local", "~133", "dev1234567890abcdef... (64 chars)", "✅ Sì", "🔴 Alto", "Chiave crittografia n8n (64 caratteri)"),
    ("n8n Workflow", "MEEPLEAI_API_URL", ".env.local", "~135", "http://api:8080", "✅ Sì", "🟢 Basso", "URL API per workflow n8n"),

    # === OBSERVABILITY ===
    ("Observability", "SEQ_URL", ".env.local", "~145", "http://seq:5341", "❌ No", "🟢 Basso", "URL Seq logging"),
    ("Observability", "PROMETHEUS_RETENTION_TIME", ".env.local", "~146", "7d", "❌ No", "🟢 Basso", "Retention metriche Prometheus"),
    ("Observability", "GF_SECURITY_ADMIN_USER", ".env.local", "~150", "admin", "❌ No", "🟡 Medio", "Username admin Grafana"),
    ("Observability", "GF_SECURITY_ADMIN_PASSWORD", ".env.local", "~151", "admin", "❌ No", "🔴 Alto", "Password admin Grafana"),
    ("Observability", "GF_SERVER_ROOT_URL", ".env.local", "~152", "http://localhost:3001", "❌ No", "🟢 Basso", "URL root Grafana"),
    ("Observability", "HYPERDX_API_KEY", "infra/env/api.env.dev", "~20", "7da70858-4d51-442f-aa4e-7ca3170f1b2f", "❌ No", "🟡 Medio", "API Key HyperDX"),
    ("Observability", "OTLP_ENDPOINT", "infra/env/api.env.dev", "~21", "http://otel-collector:4318", "❌ No", "🟢 Basso", "Endpoint OpenTelemetry"),
    ("Observability", "LOG_LEVEL", ".env.local", "~160", "Debug / Information / Warning", "❌ No", "🟢 Basso", "Livello log"),
    ("Observability", "ENABLE_STRUCTURED_LOGGING", ".env.local", "~161", "true", "❌ No", "🟢 Basso", "Abilita log strutturati JSON"),
    ("Observability", "ENABLE_METRICS", ".env.local", "~162", "true", "❌ No", "🟢 Basso", "Abilita metriche Prometheus"),

    # === EMAIL ===
    ("Email", "Email__SmtpHost", ".env.local", "~170", "mailpit (dev) / smtp.gmail.com (prod)", "❌ No", "🟢 Basso", "Host SMTP"),
    ("Email", "Email__SmtpPort", ".env.local", "~171", "1025 (dev) / 587 (prod)", "❌ No", "🟢 Basso", "Porta SMTP"),
    ("Email", "Email__EnableSsl", ".env.local", "~172", "false (dev) / true (prod)", "❌ No", "🟢 Basso", "Abilita SSL/TLS"),
    ("Email", "Email__SmtpUsername", ".env.local", "~173", "(vuoto dev) / user@gmail.com", "❌ No", "🟡 Medio", "Username SMTP"),
    ("Email", "Email__SmtpPassword", ".env.local", "~174", "(vuoto dev) / app-password", "❌ No", "🔴 Alto", "Password SMTP (App Password per Gmail)"),
    ("Email", "Email__FromAddress", ".env.local", "~175", "noreply@meepleai.dev", "❌ No", "🟢 Basso", "Indirizzo mittente"),
    ("Email", "Email__FromName", ".env.local", "~176", "MeepleAI Dev", "❌ No", "🟢 Basso", "Nome mittente"),

    # === ALERTING ===
    ("Alerting", "ALERTING_ENABLED", ".env.local", "~185", "false (dev) / true (prod)", "❌ No", "🟢 Basso", "Abilita alerting"),
    ("Alerting", "ALERTING_EMAIL_ENABLED", ".env.local", "~186", "false", "❌ No", "🟢 Basso", "Abilita alert via email"),
    ("Alerting", "ALERTING_SLACK_ENABLED", ".env.local", "~187", "false", "❌ No", "🟢 Basso", "Abilita alert via Slack"),
    ("Alerting", "ALERTING_PAGERDUTY_ENABLED", ".env.local", "~188", "false", "❌ No", "🟢 Basso", "Abilita alert via PagerDuty"),
    ("Alerting", "SLACK_WEBHOOK_URL", "infra/secrets/slack-webhook-url.txt", "file", "https://hooks.slack.com/services/xxx", "❌ No", "🔴 Alto", "Webhook URL Slack"),
    ("Alerting", "PAGERDUTY_INTEGRATION_KEY", "infra/secrets/pagerduty-integration-key.txt", "file", "xxxxx", "❌ No", "🔴 Alto", "Integration key PagerDuty"),

    # === FEATURE FLAGS ===
    ("Feature Flags", "FEATURE_PROMPT_DATABASE", ".env.local", "~200", "true", "❌ No", "🟢 Basso", "Abilita database prompt"),
    ("Feature Flags", "FEATURE_STREAMING_RESPONSES", ".env.local", "~201", "true", "❌ No", "🟢 Basso", "Abilita risposte streaming SSE"),
    ("Feature Flags", "FEATURE_SETUP_GUIDE_GENERATION", ".env.local", "~202", "true", "❌ No", "🟢 Basso", "Abilita generazione guide setup"),
    ("Feature Flags", "FEATURE_CHAT_EXPORT", ".env.local", "~203", "true", "❌ No", "🟢 Basso", "Abilita export chat"),
    ("Feature Flags", "FEATURE_MESSAGE_EDIT_DELETE", ".env.local", "~204", "true", "❌ No", "🟢 Basso", "Abilita modifica/cancella messaggi"),
    ("Feature Flags", "FEATURE_PDF_UPLOAD", ".env.local", "~205", "true", "❌ No", "🟢 Basso", "Abilita upload PDF"),

    # === PERFORMANCE ===
    ("Performance", "CACHE_OPTIMIZATION_ENABLED", ".env.local", "~215", "false (dev) / true (prod)", "❌ No", "🟢 Basso", "Abilita ottimizzazione cache"),
    ("Performance", "CACHE_WARMING_ENABLED", ".env.local", "~216", "false (dev) / true (prod)", "❌ No", "🟢 Basso", "Abilita cache warming"),
    ("Performance", "FOLLOWUP_QUESTIONS_ENABLED", ".env.local", "~220", "true", "❌ No", "🟢 Basso", "Abilita domande follow-up"),
    ("Performance", "FOLLOWUP_QUESTIONS_MAX_PER_RESPONSE", ".env.local", "~221", "3", "❌ No", "🟢 Basso", "Max domande follow-up"),
    ("Performance", "FOLLOWUP_QUESTIONS_TIMEOUT_MS", ".env.local", "~222", "5000", "❌ No", "🟢 Basso", "Timeout generazione follow-up (ms)"),
    ("Performance", "QUALITY_EVALUATION_ENABLED", ".env.local", "~225", "false", "❌ No", "🟢 Basso", "Abilita valutazione qualità risposte"),
    ("Performance", "RAG_EVALUATION_ENABLED", ".env.local", "~226", "false", "❌ No", "🟢 Basso", "Abilita valutazione RAG"),

    # === EXTERNAL APIs ===
    ("External APIs", "BGG_API_TOKEN", "infra/env/api.env.dev", "~25", "fbf56d09-385a-43fc-985f-305dbed536c9", "❌ No", "🟡 Medio", "Token API BoardGameGeek"),
    ("External APIs", "SENTRY_DSN", ".env.local", "~235", "https://xxx@sentry.io/xxx", "❌ No", "🟡 Medio", "DSN Sentry"),
    ("External APIs", "SENTRY_AUTH_TOKEN", ".env.local", "~236", "sntrys_xxxxx", "❌ No", "🔴 Alto", "Auth token Sentry"),

    # === DOCKER ===
    ("Docker", "COMPOSE_PROJECT_NAME", ".env.local", "~245", "meepleai-dev", "❌ No", "🟢 Basso", "Nome progetto Docker Compose"),
    ("Docker", "DOTNET_USE_POLLING_FILE_WATCHER", ".env.local", "~246", "true", "❌ No", "🟢 Basso", "File watcher polling (Docker volume)"),
    ("Docker", "DOTNET_WATCH_SUPPRESS_LAUNCH_BROWSER", ".env.local", "~247", "true", "❌ No", "🟢 Basso", "Non aprire browser auto"),
]


def create_excel():
    """Crea file Excel con le variabili di configurazione."""

    wb = Workbook()
    ws = wb.active
    ws.title = "Variabili Configurazione"

    # Headers
    headers = [
        "Categoria",
        "Variabile",
        "File Sorgente",
        "Posizione",
        "Esempio Valore",
        "Obbligatorio",
        "Sicurezza",
        "Note"
    ]

    # Larghezze colonne
    column_widths = [22, 40, 30, 15, 55, 18, 15, 50]

    for col_num, (header, width) in enumerate(zip(headers, column_widths), 1):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = width
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Dati
    current_category = None
    for row_num, row_data in enumerate(CONFIG_DATA, 2):
        category = row_data[0]

        # Colora riga categoria
        is_new_category = category != current_category
        current_category = category

        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical='center', wrap_text=True)

            # Colora prima cella di nuova categoria
            if col_num == 1 and is_new_category:
                cell.fill = CATEGORY_FILL
                cell.font = CATEGORY_FONT

            # Colora colonna Obbligatorio
            if col_num == 6:
                if "✅" in str(value):
                    cell.fill = REQUIRED_FILL
                elif "❌" in str(value):
                    cell.fill = OPTIONAL_FILL

            # Colora colonna Sicurezza
            if col_num == 7:
                if "🔴" in str(value):
                    cell.fill = HIGH_SECURITY_FILL

    # Altezza righe
    for row in range(1, len(CONFIG_DATA) + 2):
        ws.row_dimensions[row].height = 22

    # === SECONDO FOGLIO: Riepilogo ===
    ws2 = wb.create_sheet("Riepilogo")

    summary_data = [
        ["RIEPILOGO CONFIGURAZIONE MEEPLEAI", ""],
        ["", ""],
        ["Statistiche", ""],
        ["Totale Variabili", len(CONFIG_DATA)],
        ["Variabili Obbligatorie", sum(1 for d in CONFIG_DATA if "✅" in d[5] and "Dev" not in d[5])],
        ["Variabili Alta Sicurezza", sum(1 for d in CONFIG_DATA if "🔴" in d[6])],
        ["", ""],
        ["File Principali", ""],
        [".env.local", "Variabili ambiente locali (da copiare da .env.development.example)"],
        ["appsettings.json", "Configurazione .NET (apps/api/src/Api/)"],
        ["docker-compose.yml", "Orchestrazione Docker (infra/)"],
        ["infra/secrets/", "Docker secrets per produzione"],
        ["", ""],
        ["Variabili CRITICHE da configurare", ""],
        ["OPENROUTER_API_KEY", "Essenziale per funzionalità AI - ottieni da openrouter.ai"],
        ["POSTGRES_PASSWORD", "Cambiare in produzione!"],
        ["INITIAL_ADMIN_PASSWORD", "Cambiare in produzione!"],
        ["N8N_ENCRYPTION_KEY", "64 caratteri - generare con: openssl rand -hex 32"],
        ["", ""],
        ["Link Utili", ""],
        ["OpenRouter API Keys", "https://openrouter.ai/keys"],
        ["Google OAuth Console", "https://console.cloud.google.com/apis/credentials"],
        ["Discord Developer Portal", "https://discord.com/developers/applications"],
        ["GitHub OAuth Apps", "https://github.com/settings/developers"],
    ]

    for row_num, row_data in enumerate(summary_data, 1):
        for col_num, value in enumerate(row_data, 1):
            cell = ws2.cell(row=row_num, column=col_num, value=value)
            if row_num in [1, 3, 8, 14, 20]:
                cell.font = Font(bold=True, size=12)
            if row_num == 1:
                cell.font = Font(bold=True, size=14)

    ws2.column_dimensions['A'].width = 35
    ws2.column_dimensions['B'].width = 70

    # === TERZO FOGLIO: Quick Start ===
    ws3 = wb.create_sheet("Quick Start")

    quickstart_data = [
        ["QUICK START - Configurazione Minima", ""],
        ["", ""],
        ["Passo", "Comando / Azione"],
        ["1. Copia template", "cp .env.development.example .env.local"],
        ["2. Configura API Key AI", "Modifica OPENROUTER_API_KEY in .env.local"],
        ["3. Avvia infrastruttura", "cd infra && docker compose up -d postgres qdrant redis"],
        ["4. Avvia backend", "cd apps/api/src/Api && dotnet run"],
        ["5. Avvia frontend", "cd apps/web && pnpm dev"],
        ["6. Verifica", "Apri http://localhost:3000"],
        ["", ""],
        ["Valori Minimi Richiesti", ""],
        ["OPENROUTER_API_KEY", "Ottieni da https://openrouter.ai/keys"],
        ["", ""],
        ["Tutti gli altri valori hanno default funzionanti per sviluppo locale.", ""],
    ]

    for row_num, row_data in enumerate(quickstart_data, 1):
        for col_num, value in enumerate(row_data, 1):
            cell = ws3.cell(row=row_num, column=col_num, value=value)
            if row_num in [1, 3, 11]:
                cell.font = Font(bold=True, size=12)
            if row_num == 1:
                cell.font = Font(bold=True, size=14)

    ws3.column_dimensions['A'].width = 30
    ws3.column_dimensions['B'].width = 70

    # Salva
    output_path = Path(__file__).parent.parent.parent / "docs" / "02-development" / "MeepleAI_Configuration_Variables.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    print(f"✅ File Excel creato: {output_path}")
    print(f"   - Foglio 1: {len(CONFIG_DATA)} variabili di configurazione")
    print(f"   - Foglio 2: Riepilogo e statistiche")
    print(f"   - Foglio 3: Quick Start")

    return output_path


if __name__ == "__main__":
    create_excel()
